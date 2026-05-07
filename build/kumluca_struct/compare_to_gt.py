"""Yapisal pipeline ciktisi vs Kumluca ground truth karsilastirmasi."""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from metraj.core.structural import StructuralPipeline  # noqa: E402

DWG = ROOT / "ornekRef" / "kumluca kaba ataşman na.dwg"
GT = ROOT / "ornekRef" / "kumluca kaba.xlsx"


def normalize(s: str) -> str:
    if not isinstance(s, str):
        return ""
    tr = str.maketrans({"İ": "I", "I": "I", "ı": "I", "i": "I",
                        "Ö": "O", "ö": "O", "Ü": "U", "ü": "U",
                        "Ş": "S", "ş": "S", "Ç": "C", "ç": "C",
                        "Ğ": "G", "ğ": "G"})
    return s.translate(tr).upper()


def categorize(label: str) -> str | None:
    s = normalize(label)
    if "MINHA" in s:
        return "MINHA"
    if "PARAPET" in s:
        return "PARAPET"
    if "GRO" in s and "BETON" in s:
        return "GROBETON"
    if "ASANSOR" in s:
        return "ASANSOR"
    if "BACA" in s:
        return "BACA"
    if "TEMEL" in s:
        return "TEMEL"
    if "KOLON" in s and "YERLERI" not in s:
        return "KOLON"
    if "PERDE" in s:
        return "PERDE"
    if "KIRIS" in s:
        return "KIRIS"
    if "DOSEME" in s and "MINHA" not in s and "YAN" not in s and "BOSLUK" not in s:
        return "DOSEME"
    if "DOSEME" in s and "YAN" in s:
        return "DOSEME YAN"
    return None


def read_gt():
    wb = openpyxl.load_workbook(GT, data_only=True)
    kalip = defaultdict(float)
    for r in wb["A KALIP"].iter_rows(values_only=True):
        label = r[5] if len(r) > 5 else None
        toplam = r[8] if len(r) > 8 else None
        if not isinstance(toplam, (int, float)) or label is None:
            continue
        if "TOPLAM" in normalize(label):
            continue
        cat = categorize(label)
        if cat:
            kalip[cat] += toplam
    beton = defaultdict(float)
    for r in wb["A  BETON"].iter_rows(values_only=True):
        label = r[5] if len(r) > 5 else None
        toplam = r[8] if len(r) > 8 else None
        if not isinstance(toplam, (int, float)) or label is None:
            continue
        cat = categorize(label)
        if cat:
            beton[cat] += toplam
    return kalip, beton


def categorize_calcrow(category: str, label: str) -> str | None:
    """CalcRow.category ve label'dan ortak kategori cikar."""
    s = normalize(label)
    if "MINHA" in s:
        return "MINHA"
    return categorize(category) or categorize(label)


def main():
    print("=== Kumluca DWG yapisal pipeline ciktisi ===\n")
    p = StructuralPipeline()
    res = p.run(DWG, output_dir=ROOT / "build" / "kumluca_struct", write_excel=True)

    print(f"Plan kumesi: {res.plan_count}")
    print(f"Sistem KALIP toplami : {res.report.formwork_total_m2:.2f} m2")
    print(f"Sistem BETON toplami : {res.report.concrete_total_m3:.2f} m3")
    print()

    sys_kalip = defaultdict(float)
    sys_beton = defaultdict(float)
    for r in res.report.formwork_rows:
        cat = categorize_calcrow(r.category, r.label)
        if cat:
            sys_kalip[cat] += r.total
    for r in res.report.concrete_rows:
        cat = categorize_calcrow(r.category, r.label)
        if cat:
            sys_beton[cat] += r.total

    gt_kalip, gt_beton = read_gt()
    print(f"Ground truth KALIP toplami: {sum(gt_kalip.values()):.2f} m2")
    print(f"Ground truth BETON toplami: {sum(gt_beton.values()):.2f} m3")
    print()

    print("=== KALIP (m2) - kategori karsilastirmasi ===")
    cats = sorted(set(list(sys_kalip.keys()) + list(gt_kalip.keys())))
    print(f"{'Kategori':<14} {'Sistem':>10} {'GT':>10} {'Fark':>10} {'Sapma %':>10}")
    for c in cats:
        s, g = sys_kalip[c], gt_kalip[c]
        diff = s - g
        pct = ((s - g) / g * 100.0) if g else float("inf")
        pct_str = f"{pct:>9.1f}%" if g else "       N/A"
        print(f"{c:<14} {s:>10.1f} {g:>10.1f} {diff:>10.1f} {pct_str}")
    s_tot, g_tot = sum(sys_kalip.values()), sum(gt_kalip.values())
    diff = s_tot - g_tot
    pct = (diff / g_tot * 100.0) if g_tot else 0
    print(f"{'TOPLAM':<14} {s_tot:>10.1f} {g_tot:>10.1f} {diff:>10.1f} {pct:>9.1f}%")

    print()
    print("=== BETON (m3) - kategori karsilastirmasi ===")
    cats = sorted(set(list(sys_beton.keys()) + list(gt_beton.keys())))
    print(f"{'Kategori':<14} {'Sistem':>10} {'GT':>10} {'Fark':>10} {'Sapma %':>10}")
    for c in cats:
        s, g = sys_beton[c], gt_beton[c]
        diff = s - g
        pct = ((s - g) / g * 100.0) if g else float("inf")
        pct_str = f"{pct:>9.1f}%" if g else "       N/A"
        print(f"{c:<14} {s:>10.1f} {g:>10.1f} {diff:>10.1f} {pct_str}")
    s_tot, g_tot = sum(sys_beton.values()), sum(gt_beton.values())
    diff = s_tot - g_tot
    pct = (diff / g_tot * 100.0) if g_tot else 0
    print(f"{'TOPLAM':<14} {s_tot:>10.1f} {g_tot:>10.1f} {diff:>10.1f} {pct:>9.1f}%")


if __name__ == "__main__":
    main()
