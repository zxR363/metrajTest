"""Yapisal (kaba insaat) metraj karsilastirma scripti.

Kumluca DWG'si yapisal (kalip + beton) metraj icin hazirlandi; bizim
sistem ise mimari metraj uretiyor. Bu script DWG'den dogrudan okunan
geometrik bilgilerle ground truth Excel'i karsilastirir.

Hesaplananlar (her ikisi de DWG'de):
- KOLON NA      -> kapali polyline toplam alani (m^2)
- KIRIS NA      -> kapali polyline toplam alani (m^2)  + cevre uzunlugu
- PERDE NA + ASANSOR KULE PERDE NA -> kapali polyline alan + cevre
- DOSEME NA     -> kapali polyline toplam alani (m^2)
- DOSEME MINHA NA -> kapali polyline toplam alan (m^2, dusulecek)
- 30/50/110 CM PARAPET NA -> polyline uzunlugu
- GROBETON NA  -> kapali polyline alani

Sonra bu sayilar Kumluca ground truth Excel'iyle yan yana yazilir.
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from shapely.geometry import Polygon  # noqa: E402
from shapely.ops import unary_union  # noqa: E402

from metraj.core.cad_io.converter import DwgConverter  # noqa: E402
from metraj.core.cad_io.dxf_reader import DxfReader, RawCadModel  # noqa: E402

DWG = ROOT / "ornekRef" / "kumluca kaba ataşman na.dwg"


def _polyline_area(pts):
    if len(pts) < 3:
        return 0.0
    return Polygon(pts).area if Polygon(pts).is_valid else 0.0


def _polyline_length(pts, closed):
    import math
    total = 0.0
    for (x1, y1), (x2, y2) in zip(pts, pts[1:]):
        total += math.hypot(x2 - x1, y2 - y1)
    if closed and len(pts) >= 2:
        x1, y1 = pts[-1]
        x2, y2 = pts[0]
        total += math.hypot(x2 - x1, y2 - y1)
    return total


def collect_layer_geometry(model: RawCadModel, layers: list[str]):
    """Verilen katmanlardaki kapali polylinelerin alan + cevre toplami."""
    target = {l.upper() for l in layers}
    polys = []
    open_lengths = []
    for poly in model.polylines:
        if poly.layer.upper() not in target:
            continue
        if poly.closed and len(poly.points) >= 3:
            try:
                p = Polygon(poly.points)
                if p.is_valid and p.area > 0:
                    polys.append(p)
            except Exception:
                continue
        else:
            open_lengths.append(_polyline_length(poly.points, poly.closed))

    if polys:
        merged = unary_union(polys)
        # area: total + boundary length (+inner boundaries) for forms
        if merged.geom_type == "Polygon":
            total_area = merged.area
            total_perim = merged.length
            count = 1
        else:  # MultiPolygon
            total_area = sum(p.area for p in merged.geoms)
            total_perim = sum(p.length for p in merged.geoms)
            count = len(merged.geoms)
    else:
        total_area = 0.0
        total_perim = 0.0
        count = 0
    return {
        "polygon_count": count,
        "raw_polygon_count": len(polys),
        "total_area_m2": total_area,
        "total_perimeter_m": total_perim,
        "open_polyline_total_length_m": sum(open_lengths),
    }


def main():
    print("== Kumluca Kaba Insaat Karsilastirma ==\n")
    print(f"DWG: {DWG.name}\n")

    converter = DwgConverter()
    if not converter.is_available:
        print("ODA File Converter bulunamadi.")
        sys.exit(2)
    dxf = converter.ensure_dxf(DWG)
    print(f"DXF: {dxf.name}\n")

    model = DxfReader().read(dxf)
    print(f"Toplam katman: {len(model.layers)}")
    print(f"  lines={len(model.lines)}  polylines={len(model.polylines)}  texts={len(model.texts)}\n")

    # -- Yapisal katman gruplari ----------------------------------------
    groups = {
        "KOLON":          ["KOLON NA"],
        "PERDE":          ["PERDE NA", "ASANSÖR KULE PERDE NA"],
        "KIRIS":          ["KİRİŞ NA"],
        "DOSEME":         ["DÖŞEME NA"],
        "DOSEME MINHA":   ["DÖŞEME MİNHA NA"],
        "GROBETON":       ["GROBETON NA"],
        "PARAPET 30 CM":  ["30 CM PARAPET NA"],
        "PARAPET 50 CM":  ["50 CM PARAPET NA"],
        "PARAPET 110 CM": ["110 CM PARAPET NA"],
        "BACA":           ["90 CM BACA PARAPET NA", "BACA"],
    }

    available_upper = {l.upper(): l for l in model.layers}
    results = {}
    for name, layers in groups.items():
        present = [l for l in layers if l.upper() in available_upper]
        missing = [l for l in layers if l.upper() not in available_upper]
        info = collect_layer_geometry(model, present) if present else None
        results[name] = {"layers_present": present, "layers_missing": missing, "info": info}

    print(f"{'Grup':<18} {'Katman':<35} {'Kapali#':<8} {'Alan(m2)':<10} {'Cevre(m)':<10} {'AcikUzun(m)':<10}")
    print("-" * 95)
    for name, r in results.items():
        layers_text = ", ".join(r["layers_present"]) or "(katman yok)"
        if r["info"]:
            i = r["info"]
            print(f"{name:<18} {layers_text:<35} {i['raw_polygon_count']:<8} "
                  f"{i['total_area_m2']:<10.2f} {i['total_perimeter_m']:<10.2f} "
                  f"{i['open_polyline_total_length_m']:<10.2f}")
        else:
            print(f"{name:<18} {layers_text:<35} -")

    # -- Ground truth ozeti --------------------------------------------
    print()
    print("== Ground Truth (kumluca kaba.xlsx) ==")
    print()

    import openpyxl
    gt = openpyxl.load_workbook(ROOT / "ornekRef" / "kumluca kaba.xlsx", data_only=True)

    def normalize(s: str) -> str:
        """Turkce karakterleri ASCII'ye dondur ve buyuk harfe getir."""
        if not isinstance(s, str):
            return ""
        tr = str.maketrans({"İ": "I", "I": "I", "ı": "I", "i": "I",
                            "Ö": "O", "ö": "O", "Ü": "U", "ü": "U",
                            "Ş": "S", "ş": "S", "Ç": "C", "ç": "C",
                            "Ğ": "G", "ğ": "G"})
        return s.translate(tr).upper()

    def categorize(label: str) -> str | None:
        s = normalize(label)
        if "MINHA" in s:
            return "MINHA (toplam)"
        if "PARAPET" in s and "ASANSOR" not in s:
            return "PARAPET"
        if "GRO" in s and "BETON" in s:
            return "GROBETON"
        if "ASANSOR" in s or "BACA" in s:
            return "ASANSOR/BACA"
        if "TEMEL" in s:
            return "TEMEL"
        if "KOLON" in s:
            return "KOLON"
        if "PERDE" in s:
            return "PERDE"
        if "KIRIS" in s:
            return "KIRIS"
        if "DOSEME" in s:
            return "DOSEME"
        return None

    # KALIP toplami
    sh_kalip = gt["A KALIP"]
    kalip_total = None
    kalip_categories = defaultdict(float)
    for row in sh_kalip.iter_rows(values_only=True):
        label = row[5]
        toplam = row[8]
        if isinstance(label, str) and "TOPLAM" in normalize(label):
            if isinstance(toplam, (int, float)):
                kalip_total = toplam
            continue
        if label is None or toplam is None or not isinstance(toplam, (int, float)):
            continue
        cat = categorize(label)
        if cat:
            kalip_categories[cat] += toplam

    if kalip_total is None:
        kalip_total = sum(kalip_categories.values())
    print("KALIP m2 ozeti (ground truth):")
    for k, v in kalip_categories.items():
        print(f"  {k:<20} {v:>10.2f}")
    print(f"  {'TOPLAM (hesaplanan)':<20} {kalip_total:>10.2f}")

    # BETON
    sh_beton = gt["A  BETON"]
    beton_total = None
    beton_categories = defaultdict(float)
    extra_lean = defaultdict(float)  # GRO/KORUMA/CATI degerleri
    for row in sh_beton.iter_rows(values_only=True):
        label_main = row[5] if len(row) > 5 else None
        label_summary = row[7] if len(row) > 7 else None
        toplam = row[8] if len(row) > 8 else None
        # Toplam satiri: row[7] = "C35 TOPLAM"
        if isinstance(label_summary, str) and "TOPLAM" in normalize(label_summary):
            if isinstance(toplam, (int, float)):
                beton_total = toplam
            continue
        # GRO/KORUMA/CATI satirlari row[7]'de label
        if label_main is None and label_summary and isinstance(toplam, (int, float)):
            extra_lean[normalize(label_summary)] += toplam
            continue
        if label_main is None or toplam is None or not isinstance(toplam, (int, float)):
            continue
        cat = categorize(label_main)
        if cat:
            beton_categories[cat] += toplam

    if beton_total is None:
        beton_total = sum(beton_categories.values())
    print()
    print("BETON m3 ozeti (ground truth):")
    for k, v in beton_categories.items():
        print(f"  {k:<20} {v:>10.2f}")
    print(f"  {'TOPLAM C35':<20} {beton_total:>10.2f}")
    if extra_lean:
        print("Ek beton kategorileri (kuru hesap):")
        for k, v in extra_lean.items():
            print(f"  {k:<20} {v:>10.2f}")

    # -- Karsilastirma ------------------------------------------------
    print()
    print("== Sistem cikti karsilastirmasi ==")
    print()
    print("Bizim sistem (mimari odakli) bu DWG'den 0 mahal, 0 m2 doseme")
    print("kaplamasi cikariyor cunku DWG'de mahal etiketi yok.")
    print()
    print("DWG'den dogrudan okudugumuz yapisal alanlar (yukaridaki tablo)")
    print("ground truth ile asagidaki gibi karsilastirilabilir:")
    print()

    # Ground truth genel parametreler
    h_floor = 2.85         # tipik kat-arasi temiz yukseklik
    n_floors = 6           # 0 / 3 / 6 / 9 / 12 / 15 -> 6 kat
    n_storey_columns = 6   # (-3,00)/0,00 + 5 kat-arasi
    slab_thickness = 0.15

    # DWG'deki cizimlerde her kat ayri ayri cizilmis olabilir (cogu kez 6 plan
    # yan yana) — biz 280 KOLON polygonunu okuyoruz; eger her kat ~47 kolon
    # iceriyorsa, kolon perimetresi 840/6 = 140 m gibi olur.

    def per_storey(group_name):
        r = results.get(group_name)
        if not r or not r["info"]:
            return None
        i = r["info"]
        return {
            "area_total_m2":  i["total_area_m2"] / n_floors,
            "perim_total_m":  i["total_perimeter_m"] / n_floors,
            "area_raw":       i["total_area_m2"],
            "perim_raw":      i["total_perimeter_m"],
            "polygon_count":  i["raw_polygon_count"],
            "polygon_per_floor": i["raw_polygon_count"] / n_floors,
        }

    print()
    print(f"DWG geometri hipotezi: {n_floors} kat planinin tamami ayni dosyada")
    print(f"yan yana cizilmis. Bir kat icin tahminler:")
    print()
    print(f"  {'Grup':<12} {'kat_perim':>10} {'kat_alan':>10} {'kat_polig':>10}")
    for g in ["KOLON", "PERDE", "KIRIS", "DOSEME"]:
        ps = per_storey(g)
        if ps:
            print(f"  {g:<12} {ps['perim_total_m']:>10.2f} {ps['area_total_m2']:>10.2f} {ps['polygon_per_floor']:>10.1f}")

    # -- Tahmini kalip/beton hesabi ----------------------------------
    print()
    print("== Tahmini kalip/beton (DWG geometrisinden + kat tekrari) ==")

    def estimate_formwork(group_name, height, n_repeat, twosided=False, exclude_top_bot=False):
        ps = per_storey(group_name)
        if not ps:
            return 0.0, 0.0, 0.0
        # Tek kat icin perimetre x yukseklik (kalip iki yuz icin x2)
        side = ps["perim_total_m"] * height
        if twosided:
            side *= 2
        return side, side * n_repeat, ps["area_total_m2"] * height * n_repeat

    print()
    print(f"  {'Grup':<14} {'Tahmin kalip(m2)':>18} {'GT kalip(m2)':>14} {'Tahmin beton(m3)':>18} {'GT beton(m3)':>14}")

    # KOLON: perimetre x H x kat
    ps = per_storey("KOLON")
    if ps:
        kal_kolon = ps["perim_total_m"] * h_floor * n_storey_columns
        bet_kolon = ps["area_total_m2"] * h_floor * n_storey_columns
        print(f"  {'KOLON':<14} {kal_kolon:>18.1f} {kalip_categories['KOLON']:>14.1f} "
              f"{bet_kolon:>18.1f} {beton_categories['KOLON']:>14.1f}")

    # PERDE: perimetre x H x kat (iki yuz)
    ps = per_storey("PERDE")
    if ps:
        kal_perde = ps["perim_total_m"] * h_floor * n_storey_columns  # tek yuz - icini de kapsar
        bet_perde = ps["area_total_m2"] * h_floor * n_storey_columns
        print(f"  {'PERDE':<14} {kal_perde:>18.1f} {kalip_categories['PERDE']:>14.1f} "
              f"{bet_perde:>18.1f} {beton_categories['PERDE']:>14.1f}")

    # KIRIS: GT kalip kiris = uzunluk x 0.45 (kiris tabani+yan), beton = alan x 0.45
    ps = per_storey("KIRIS")
    if ps:
        kal_kiris = ps["perim_total_m"] * 0.45 * n_floors  # kaba: cevre x kiris derinligi
        bet_kiris = ps["area_total_m2"] * 0.45 * n_floors
        print(f"  {'KIRIS':<14} {kal_kiris:>18.1f} {kalip_categories['KIRIS']:>14.1f} "
              f"{bet_kiris:>18.1f} {beton_categories['KIRIS']:>14.1f}")

    # DOSEME: kalip = alan x kat, beton = alan x 0.15 x kat
    ps = per_storey("DOSEME")
    if ps:
        kal_dos = ps["area_total_m2"] * n_floors
        bet_dos = ps["area_total_m2"] * slab_thickness * n_floors
        print(f"  {'DOSEME':<14} {kal_dos:>18.1f} {kalip_categories['DOSEME']:>14.1f} "
              f"{bet_dos:>18.1f} {beton_categories['DOSEME']:>14.1f}")

    # GROBETON
    ps = per_storey("GROBETON")
    if ps:
        kal_gro = ps["area_total_m2"]  # tek kez (zemin)
        print(f"  {'GROBETON':<14} {kal_gro:>18.1f} {kalip_categories['GROBETON']:>14.1f} "
              f"{'-':>18} {'-':>14}")

    print()
    print("Notlar:")
    print(" - Tahminlerde 6 kat tekrari ve 2.85 m kat yuksekligi varsayildi.")
    print(" - GT kalip KOLON'a (-3,00)/0,00 (h=2.85), 0/3, 3/6, 6/9, 9/12, 12/15 dahildir.")
    print(" - Bu sistemin yapisal kategori cikarmasi heuristik; gercek kalip/beton")
    print("   metraj icin ayri bir 'kaba_insaat' moduluyle kat ayrimi + minhalar")
    print("   eklenmelidir.")


if __name__ == "__main__":
    main()
