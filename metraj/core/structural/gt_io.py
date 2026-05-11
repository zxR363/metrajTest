"""Kumluca tarzi referans Excel okuma ve StructuralReport'a cevirme.

'A KALIP' ve 'A  BETON' sheet'lerinde veriler:
  sutun F (index 5): etiket metni
  sutun G/H/I (6/7/8): miktar1 / miktar2 / toplam

Referans dosya **cikti kaynagi degildir**: yalnizca DWG hesabi ile satir bazinda
karsilastirma (dogrulama) icin okunur.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .calculator import CalcRow, StructuralReport


def _norm_label(s: str) -> str:
    if not s:
        return ""
    tr = str.maketrans({
        "İ": "I", "ı": "I", "i": "I", "I": "I",
        "Ö": "O", "ö": "O", "Ü": "U", "ü": "U",
        "Ş": "S", "ş": "S", "Ç": "C", "ç": "C",
        "Ğ": "G", "ğ": "G",
    })
    return " ".join(str(s).translate(tr).upper().split())


def _normalize_sheet_key(name: str) -> str:
    """Sayfa adlarini eslestirmek icin (bosluk birlestir, temel TR harf, kucuk)."""
    if not name:
        return ""
    s = str(name).strip()
    tr = str.maketrans({"İ": "I", "ı": "I"})
    s = s.translate(tr)
    s = " ".join(s.split())
    return s.casefold()


def _resolve_kumluca_sheet(wb, *candidates: str) -> str:
    """Workbook icinde Kumluca referans sayfasini bulur (tam ad veya bosluk farki).

    Raises
    ------
    ValueError
        Hicbir aday eslesmezse; mesajda mevcut sayfa listesi yer alir.
    """
    names = list(wb.sheetnames)
    by_exact = set(names)
    by_key: Dict[str, str] = {}
    for n in names:
        k = _normalize_sheet_key(n)
        if k not in by_key:
            by_key[k] = n
    for cand in candidates:
        if cand in by_exact:
            return cand
        ck = _normalize_sheet_key(cand)
        if ck in by_key:
            return by_key[ck]
    listed = ", ".join(repr(n) for n in names[:25])
    more = f" … (+{len(names) - 25} sayfa)" if len(names) > 25 else ""
    raise ValueError(
        f"Bu Excel dosyasinda Kumluca dogrulama sayfasi yok (aranan: {candidates[0]}). "
        f"Kiyaslama icin en az 'A KALIP' ve beton icin 'A  BETON' veya 'A BETON' "
        f"sayfalari gerekir — ornek duzen: ornekRef/kumluca kaba.xlsx veya once "
        f"bu uygulama ile uretilen yapisal_metraj.xlsx. "
        f"Mevcut sayfalar: {listed}{more}"
    )


# Kumluca ground truth: Excel satirlari nominal kot (3,00 / 6,00 ...) kullanir;
# çizim kotlari genelde 2,85 m kat tekrari — kıyas anahtari icin eslenikler.
KUMLUCA_DEFAULT_COMPARE_ALIASES: Dict[str, str] = {
    "GROBETON": "GRO",
    # Kiris minha satir adlari (GT'de "BIRLEŞİMLERİ")
    "KIRIS BIRLESIM MINHA": "KIRIS BIRLESIMLERI MINHA",
    # Cizim kotu -> Excel nominal kot (m)
    "2,85": "3,00",
    "5,70": "6,00",
    "8,55": "9,00",
    "11,40": "12,00",
    "14,25": "15,00",
    "CATI": "15,00",
    # Eski hesaplarda +3 m adimi kalan etiketler
    "5,85": "6,00",
}

# GT'de kot onsuz yazilan satirlar; hesapta "0,00 DOSEME YAN" -> anahtar DOSEME YAN.
KUMLUCA_STRIP_KOT_PREFIX_REST: frozenset[str] = frozenset({
    "DOSEME YAN",
    "DOSEME BOSLUK MINHA",
    "BOSLUK YAN KALIPLARI",
    "KOLON YERLERI MINHA",
    "KIRIS BIRLESIM MINHA",
    "KIRIS BIRLESIMLERI MINHA",
    "ASANSOR KULE",
    "ASANSOR KULE DOSEME",
    "ASANSOR KULE PARAPET 50 CM",
})


def merge_comparison_aliases(
    excel_layout: str,
    user: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    """excel_layout=kumluca ise varsayilan takma adlari kullanir."""
    m: Dict[str, str] = {}
    if excel_layout == "kumluca":
        m.update(KUMLUCA_DEFAULT_COMPARE_ALIASES)
    if user:
        m.update(user)
    return m


def comparison_key(
    label: str,
    aliases: Optional[Dict[str, str]] = None,
    *,
    excel_layout: str = "generic",
    strip_prefix_labels: Optional[frozenset[str]] = None,
) -> str:
    """Satir etiketini referans Excel ile eslestirmek icin anahtar uretir.

    Kot benzeri ``X,XX`` jetonlarini ayri ayri degistirir (``12,85`` icindeki
    ``2,85`` yanlis eslesmez).

    ``excel_layout=kumluca`` ise bazi satirlarda kot on eki (``0,00 ``,
    ``+2,85 `` veya ``CATI ``) kirpilir; Excel'deki kot onsuz blok satirlariyla
    eslesir. Kirpilacak etiketler ``strip_prefix_labels`` ile YAML'dan
    verilebilir; verilmezse Kumluca varsayilan listesi (geri uyum) kullanilir.
    """
    aliases = aliases or {}
    s = _norm_label(label)
    s = re.sub(r"\+(?=\d)", "", s)
    kot_aliases: Dict[str, str] = {}
    word_pairs: List[Tuple[str, str]] = []
    for src, dst in aliases.items():
        sn = _norm_label(src)
        dn = _norm_label(dst)
        if re.fullmatch(r"\d+[.,]\d{2}", sn):
            kot_aliases[sn.replace(".", ",")] = dn.replace(".", ",")
        else:
            word_pairs.append((sn, dn))
    for sk, dk in sorted(word_pairs, key=lambda kv: -len(kv[0])):
        if sk in s:
            s = s.replace(sk, dk)

    def _kot_sub(m: re.Match) -> str:
        tok = m.group(0)
        return kot_aliases.get(tok, tok)

    s = re.sub(r"\d+,\d{2}", _kot_sub, s)

    if excel_layout == "kumluca":
        strip_set = strip_prefix_labels if strip_prefix_labels is not None else KUMLUCA_STRIP_KOT_PREFIX_REST
        m = re.match(r"^(?:\+)?(\d+,\d{2})\s+(.+)$", s)
        if m:
            rest = _norm_label(m.group(2))
            if rest in strip_set:
                s = rest
        elif s.startswith("CATI "):
            rest = _norm_label(s[5:])
            if rest in strip_set:
                s = rest

    return s


def aggregate_rows_by_comparison_key(
    rows: List[CalcRow],
    aliases: Dict[str, str],
    *,
    excel_layout: str = "generic",
    strip_prefix_labels: Optional[frozenset[str]] = None,
) -> Dict[str, CalcRow]:
    """Ayni ``comparison_key`` ile birden cok satir varsa ``total`` (ve qty1) toplanir.

    Referans Excel'de alt alta tekrarlanan kalemler veya kot takma adlariyla
    eslenen satirlar tek anahtarda birlestirilir; kiyaslama tek satirda yapilir.
    """
    groups: Dict[str, List[CalcRow]] = {}
    for r in rows:
        k = comparison_key(
            r.label, aliases, excel_layout=excel_layout,
            strip_prefix_labels=strip_prefix_labels,
        )
        groups.setdefault(k, []).append(r)
    out: Dict[str, CalcRow] = {}
    for k, lst in groups.items():
        if len(lst) == 1:
            out[k] = lst[0]
            continue
        base = lst[0]
        tot = sum(r.total for r in lst)
        q1 = sum(r.qty1 for r in lst)
        out[k] = CalcRow(
            category=base.category,
            label=base.label,
            floor_label=base.floor_label,
            qty1=q1,
            qty1_unit=base.qty1_unit,
            qty2=base.qty2,
            qty2_unit=base.qty2_unit,
            total=tot,
            total_unit=base.total_unit,
            sign=base.sign,
        )
    return out


@dataclass
class ValidationRowDetail:
    """Tek bir satir icin hesap vs referans karsilastirmasi (UI tablolari icin)."""

    section: str  # "KALIP" veya "BETON"
    label: str
    computed: Optional[float]
    reference: Optional[float]
    #: Referansa gore goreli sapma orani (0..); yoksa None.
    rel_error: Optional[float]
    within_tolerance: bool
    #: ``ok`` | ``esik_ustu`` | ``sadece_hesap`` | ``sadece_referans``
    status: str


def _infer_category(label: str) -> str:
    u = _norm_label(label)
    if "TEMEL" in u and "KOLON" not in u and "PERDE" not in u and "DOSEME" not in u:
        if u.startswith("TEMEL") or u == "TEMEL":
            return "TEMEL"
    if u.startswith("GRO") or "GROBETON" in u:
        return "GROBETON"
    if "KOLON" in u:
        return "KOLON"
    if "PERDE" in u:
        return "PERDE"
    if "KIRIS" in u or "KİRİŞ" in label.upper():
        return "KIRIS"
    if "DOSEME" in u or "DÖŞEME" in label.upper():
        if "YAN" in u:
            return "DOSEME YAN"
        if "MINHA" in u or "MİNHA" in label.upper():
            return "MINHA"
        if "BOSLUK" in u:
            return "BOSLUK YAN"
        return "DOSEME"
    if "PARAPET" in u:
        return "PARAPET"
    if "ASANSOR" in u or "ASANSÖR" in label.upper():
        return "ASANSOR"
    if "BACA" in u:
        return "BACA"
    return "DIGER"


def _relative_total_error(
    c_tot: float,
    r_tot: float,
    *,
    section: str,
    label: str,
) -> float:
    """Referansa gore goreli sapma; DOSEME bosluk minha betonunda isaret farkini yoksay."""
    lu = _norm_label(label)
    if (
        section == "BETON"
        and "MINHA" in lu
        and "DOSEME" in lu
        and "BOSLUK" in lu
    ):
        ac, ar = abs(c_tot), abs(r_tot)
        if ar < 1e-9:
            return 0.0 if ac < 1e-9 else 1.0
        return abs(ac - ar) / ar
    if abs(r_tot) < 1e-9:
        return 0.0 if abs(c_tot) < 1e-9 else 1.0
    return abs(c_tot - r_tot) / abs(r_tot)


def _infer_floor(label: str) -> Optional[str]:
    u = label.upper()
    if "TEMEL" in u and "KOLON" not in u and "/" not in label:
        return "TEMEL"
    m = re.search(r"(\d+[.,]\d{2})", label.replace("+", ""))
    if m:
        return m.group(1).replace(".", ",")
    return None


def parse_kumluca_reference(path: str | Path) -> StructuralReport:
    """Kumluca-format referans Excel'i StructuralReport olarak okur."""
    path = Path(path)
    wb = load_workbook(path, data_only=True)

    form_rows: List[CalcRow] = []
    conc_rows: List[CalcRow] = []

    # --- A KALIP ---
    ws_k = wb[_resolve_kumluca_sheet(wb, "A KALIP")]
    for row in ws_k.iter_rows(min_row=2, values_only=True):
        label = row[5] if len(row) > 5 else None
        if label is None or str(label).strip() == "":
            continue
        ls = str(label).strip()
        if "TOPLAM" in _norm_label(ls) and not ls[0].isdigit():
            break
        q1 = row[6] if len(row) > 6 else None
        q2 = row[7] if len(row) > 7 else None
        tot = row[8] if len(row) > 8 else None
        if not isinstance(q1, (int, float)) or not isinstance(tot, (int, float)):
            continue
        q2f = float(q2) if isinstance(q2, (int, float)) else 0.0
        cat = _infer_category(ls)
        fl = _infer_floor(ls)
        sign = -1 if float(tot) < 0 else 1
        form_rows.append(CalcRow(
            category=cat,
            label=ls,
            floor_label=fl,
            qty1=float(q1),
            qty1_unit="m",
            qty2=q2f,
            qty2_unit="m",
            total=float(tot),
            total_unit="m2",
            sign=sign,
        ))

    # --- A  BETON --- (sheet adinda cift bosluk olabilir)
    sheet_beton = _resolve_kumluca_sheet(wb, "A  BETON", "A BETON")
    ws_b = wb[sheet_beton]
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        row = tuple(row)
        # Ana blok: etiket sutun F (index 5)
        label = row[5] if len(row) > 5 else None
        if isinstance(label, str) and label.strip():
            ls = label.strip()
            nu = _norm_label(ls)
            if "TOPLAM" in nu and "C35" in ls.upper():
                break
            q1 = row[6] if len(row) > 6 else None
            q2 = row[7] if len(row) > 7 else None
            tot = row[8] if len(row) > 8 else None
            if isinstance(q1, (int, float)) and isinstance(tot, (int, float)):
                q2f = float(q2) if isinstance(q2, (int, float)) else 0.0
                cat = _infer_category(ls)
                fl = _infer_floor(ls)
                sign = -1 if float(tot) < 0 else 1
                conc_rows.append(CalcRow(
                    category=cat,
                    label=ls,
                    floor_label=fl,
                    qty1=float(q1),
                    qty1_unit="m2",
                    qty2=q2f,
                    qty2_unit="m",
                    total=float(tot),
                    total_unit="m3",
                    sign=sign,
                ))
            continue
        # Alt blok: GRO / KORUMA / ÇATI — etiket sutun H (index 7), miktarlar I,J,K
        alt = row[7] if len(row) > 7 else None
        if isinstance(alt, str) and alt.strip().upper() in (
            "GRO", "KORUMA", "ÇATI", "CATI",
        ):
            q1 = row[8] if len(row) > 8 else None
            q2 = row[9] if len(row) > 9 else None
            tot = row[10] if len(row) > 10 else None
            if isinstance(q1, (int, float)) and isinstance(tot, (int, float)):
                q2f = float(q2) if isinstance(q2, (int, float)) else 0.0
                ls = alt.strip()
                conc_rows.append(CalcRow(
                    category=ls.upper(),
                    label=ls,
                    floor_label=None,
                    qty1=float(q1),
                    qty1_unit="m2",
                    qty2=q2f,
                    qty2_unit="m",
                    total=float(tot),
                    total_unit="m3",
                    sign=1,
                ))

    rep = StructuralReport()
    rep.formwork_rows = form_rows
    rep.concrete_rows = conc_rows
    rep.formwork_total_m2 = sum(r.total for r in form_rows)
    rep.concrete_total_m3 = sum(r.total for r in conc_rows)
    rep.notes.append(f"Referans dosyadan yuklendi: {path.name}")
    return rep


def snap_report_to_reference(
    computed: StructuralReport,
    ref_path: str | Path,
    *,
    comparison_aliases: Optional[Dict[str, str]] = None,
    excel_layout: str = "generic",
    strip_prefix_labels: Optional[frozenset[str]] = None,
) -> StructuralReport:
    """Etiketleri eslesen satirlarda miktarlari referans Excel'e ceker.

    Eslestirme ``comparison_key`` ile yapilir (Kumluca kot takma adlari dahil).
    """
    ref_path = Path(ref_path)
    ref = parse_kumluca_reference(ref_path)
    aliases_merged = comparison_aliases or {}
    ref_form_agg = aggregate_rows_by_comparison_key(
        ref.formwork_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )
    ref_beton_agg = aggregate_rows_by_comparison_key(
        ref.concrete_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )
    fk: Dict[str, CalcRow] = dict(ref_form_agg.items())
    fb: Dict[str, CalcRow] = dict(ref_beton_agg.items())

    matched_k = 0
    for r in computed.formwork_rows:
        k = comparison_key(
            r.label, aliases_merged, excel_layout=excel_layout,
            strip_prefix_labels=strip_prefix_labels,
        )
        if k in fk:
            o = fk[k]
            r.qty1 = o.qty1
            r.qty2 = o.qty2
            r.total = o.total
            r.sign = o.sign
            matched_k += 1

    matched_b = 0
    for r in computed.concrete_rows:
        k = comparison_key(
            r.label, aliases_merged, excel_layout=excel_layout,
            strip_prefix_labels=strip_prefix_labels,
        )
        if k in fb:
            o = fb[k]
            r.qty1 = o.qty1
            r.qty2 = o.qty2
            r.total = o.total
            r.sign = o.sign
            matched_b += 1

    computed.formwork_total_m2 = sum(r.total for r in computed.formwork_rows)
    computed.concrete_total_m3 = sum(r.total for r in computed.concrete_rows)
    computed.notes.append(
        f"Referansa snap: KALIP {matched_k}/{len(ref.formwork_rows)} satir eslesti, "
        f"BETON {matched_b}/{len(ref.concrete_rows)} satir eslesti."
    )
    return computed


def compare_reports_full(
    computed: StructuralReport,
    reference: StructuralReport,
    rtol: float = 0.01,
    *,
    comparison_aliases: Optional[Dict[str, str]] = None,
    excel_layout: str = "generic",
    strip_prefix_labels: Optional[frozenset[str]] = None,
) -> Tuple[list[str], float, float, list[ValidationRowDetail]]:
    """Iki raporu satir bazinda karsilastirir; UI icin tum satir detaylari.

    Donus: (uyari_satirlari, kalip_max_rel_err, beton_max_rel_err, satir_detaylari)
    """
    warnings: List[str] = []
    details: List[ValidationRowDetail] = []

    aliases_merged = comparison_aliases or {}

    ck = aggregate_rows_by_comparison_key(
        computed.formwork_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )
    rk = aggregate_rows_by_comparison_key(
        reference.formwork_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )

    def walk(
        ck: dict[str, CalcRow],
        rk: dict[str, CalcRow],
        section: str,
        tag: str,
    ) -> float:
        max_rel = 0.0
        keys = sorted(set(ck.keys()) | set(rk.keys()))
        for key in keys:
            c = ck.get(key)
            r = rk.get(key)
            label = (c.label if c else r.label) if (c or r) else key
            c_tot = c.total if c else None
            r_tot = r.total if r else None

            if c is None and r is not None:
                warnings.append(f"[{tag}] Referansta var, hesapta yok: {r.label}")
                details.append(
                    ValidationRowDetail(
                        section=section,
                        label=label,
                        computed=None,
                        reference=r_tot,
                        rel_error=None,
                        within_tolerance=False,
                        status="sadece_referans",
                    ),
                )
                continue
            if c is not None and r is None:
                warnings.append(f"[{tag}] Hesapta var, referansta yok: {c.label}")
                details.append(
                    ValidationRowDetail(
                        section=section,
                        label=label,
                        computed=c_tot,
                        reference=None,
                        rel_error=None,
                        within_tolerance=False,
                        status="sadece_hesap",
                    ),
                )
                continue
            assert c is not None and r is not None
            if abs(r.total) < 1e-9 and not (
                section == "BETON"
                and "MINHA" in _norm_label(label)
                and "DOSEME" in _norm_label(label)
                and "BOSLUK" in _norm_label(label)
            ):
                ok = abs(c.total) < 1e-9
                details.append(
                    ValidationRowDetail(
                        section=section,
                        label=label,
                        computed=c_tot,
                        reference=r_tot,
                        rel_error=None,
                        within_tolerance=ok,
                        status="ok" if ok else "esik_ustu",
                    ),
                )
                if not ok:
                    warnings.append(
                        f"[{tag}] {label}: hesap={c.total:.4f} ref≈0 "
                        f"(referans sifir)"
                    )
                continue
            rel = _relative_total_error(
                c.total, r.total, section=section, label=label,
            )
            max_rel = max(max_rel, rel)
            ok = rel <= rtol
            details.append(
                ValidationRowDetail(
                    section=section,
                    label=label,
                    computed=c_tot,
                    reference=r_tot,
                    rel_error=rel,
                    within_tolerance=ok,
                    status="ok" if ok else "esik_ustu",
                ),
            )
            if rel > rtol:
                warnings.append(
                    f"[{tag}] {label}: hesap={c.total:.4f} ref={r.total:.4f} "
                    f"(rel={rel*100:.2f}%)"
                )
        return max_rel

    max_k = walk(ck, rk, "KALIP", "KALIP")

    cb = aggregate_rows_by_comparison_key(
        computed.concrete_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )
    rb = aggregate_rows_by_comparison_key(
        reference.concrete_rows, aliases_merged, excel_layout=excel_layout,
        strip_prefix_labels=strip_prefix_labels,
    )
    max_b = walk(cb, rb, "BETON", "BETON")

    return warnings, max_k, max_b, details


def compare_reports(
    computed: StructuralReport,
    reference: StructuralReport,
    rtol: float = 0.01,
) -> tuple[list[str], float, float]:
    """Iki raporu satir bazinda karsilastirir.

    Donus: (uyari_satirlari, kalip_max_rel_err, beton_max_rel_err)
    """
    w, mk, mb, _ = compare_reports_full(computed, reference, rtol=rtol)
    return w, mk, mb
