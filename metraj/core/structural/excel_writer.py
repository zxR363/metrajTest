"""Yapisal metraj Excel raporu (KALIP + BETON sheet'leri).

Kumluca ground truth gibi iki sheet uretir:
- 'A KALIP' : kategori, label, uzunluk, yukseklik/derinlik, toplam (m^2)
- 'A BETON': kategori, label, alan, yukseklik/kalinlik, toplam (m^3)

Sutun yerlesimi ground truth ile uyumlu (F sutunundan itibaren etiket).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .calculator import CalcRow, StructuralReport

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="305496")
_MINHA_FONT = Font(color="C00000")  # negatif satir kirmizi


def _write_rows(ws, rows: Sequence[CalcRow], qty_label: str, total_label: str) -> None:
    # Basliklar (Kumluca'daki gibi F sutunundan baslayalim ama ilk birkac sutunu
    # bos birakmak gerek mi? Pratik: A sutununa kategori, B'ye etiket, C/D/E
    # uzunluk/yukseklik/toplam.  Daha temiz okunur)
    headers = ["KAT", "KATEGORI", "ACIKLAMA", qty_label, "Y/H", total_label]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
    row_idx = 2
    for r in rows:
        ws.cell(row=row_idx, column=1, value=r.floor_label or "")
        ws.cell(row=row_idx, column=2, value=r.category)
        ws.cell(row=row_idx, column=3, value=r.label)
        ws.cell(row=row_idx, column=4, value=round(r.qty1, 3))
        ws.cell(row=row_idx, column=5, value=round(r.qty2, 3))
        ws.cell(row=row_idx, column=6, value=round(r.total, 3))
        if r.sign < 0 or r.total < 0:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).font = _MINHA_FONT
        row_idx += 1
    # Toplam
    total = sum(r.total for r in rows)
    ws.cell(row=row_idx, column=3, value="TOPLAM").font = _HEADER_FONT
    c = ws.cell(row=row_idx, column=6, value=round(total, 3))
    c.font = _TOTAL_FONT
    c.fill = _TOTAL_FILL

    # Genislikler
    widths = [12, 14, 36, 14, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_structural_xlsx(report: StructuralReport, output_path: Path,
                          project_name: str = "Yapisal Metraj") -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Ozet sheet
    ws_summary = wb.active
    ws_summary.title = "Ozet"
    ws_summary["A1"] = project_name
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Yapisal Metraj Ozeti"
    ws_summary["A3"].font = _HEADER_FONT
    ws_summary["A5"] = "Kalip Toplami"
    ws_summary["B5"] = round(report.formwork_total_m2, 2)
    ws_summary["C5"] = "m2"
    ws_summary["A6"] = "Beton Toplami"
    ws_summary["B6"] = round(report.concrete_total_m3, 2)
    ws_summary["C6"] = "m3"
    if report.notes:
        ws_summary["A8"] = "Notlar:"
        ws_summary["A8"].font = _HEADER_FONT
        for i, note in enumerate(report.notes, start=9):
            ws_summary.cell(row=i, column=1, value=note)

    # Kalip
    ws_kal = wb.create_sheet("A KALIP")
    _write_rows(ws_kal, report.formwork_rows, qty_label="UZUNLUK/CEVRE (m)", total_label="TOPLAM (m2)")

    # Beton
    ws_bet = wb.create_sheet("A  BETON")
    _write_rows(ws_bet, report.concrete_rows, qty_label="ALAN/UZUNLUK", total_label="TOPLAM (m3)")

    # Kat bazli ozet
    if report.formwork_rows:
        ws_floor = wb.create_sheet("Kat Bazli")
        ws_floor["A1"] = "Kat Bazli Ozet"
        ws_floor["A1"].font = _HEADER_FONT
        headers = ["KAT", "KALIP (m2)", "BETON (m3)"]
        for col, h in enumerate(headers, start=1):
            ws_floor.cell(row=2, column=col, value=h).font = _HEADER_FONT
        floor_kalip = {}
        floor_beton = {}
        for r in report.formwork_rows:
            floor_kalip[r.floor_label or "?"] = floor_kalip.get(r.floor_label or "?", 0) + r.total
        for r in report.concrete_rows:
            floor_beton[r.floor_label or "?"] = floor_beton.get(r.floor_label or "?", 0) + r.total
        all_floors = sorted(set(list(floor_kalip.keys()) + list(floor_beton.keys())))
        for i, fl in enumerate(all_floors, start=3):
            ws_floor.cell(row=i, column=1, value=fl)
            ws_floor.cell(row=i, column=2, value=round(floor_kalip.get(fl, 0), 2))
            ws_floor.cell(row=i, column=3, value=round(floor_beton.get(fl, 0), 2))
        for col, w in enumerate([16, 14, 14], start=1):
            ws_floor.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    logger.info("Yapisal Excel raporu yazildi: %s", output_path)
    return output_path


def write_kumluca_reference_layout(
    report: StructuralReport,
    output_path: Path,
    project_name: str = "Yapisal Metraj",
) -> Path:
    """Kumluca ground truth ile ayni sutun yerlesimi (F=etiket, G/H/I=miktarlar).

    Ilk satir basliklari GT ile uyumlu; A-E bos kalir."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Ozet"
    ws0["A1"] = project_name
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = "Dogrulanmis metraj (referans layout)"
    ws0["A5"] = "Kalip Toplami"
    ws0["B5"] = round(report.formwork_total_m2, 6)
    ws0["A6"] = "Beton Toplami"
    ws0["B6"] = round(report.concrete_total_m3, 6)

    # --- A KALIP (sutun F-I, 1-based: F=6)
    ws_k = wb.create_sheet("A KALIP")
    ws_k.cell(row=1, column=7, value="UZUNLUK")
    ws_k.cell(row=1, column=8, value="YÜKSEKLİK")
    ws_k.cell(row=1, column=9, value="TOPLAM")
    for c in (7, 8, 9):
        ws_k.cell(row=1, column=c).font = _HEADER_FONT
    ri = 2
    for r in report.formwork_rows:
        ws_k.cell(row=ri, column=6, value=r.label)
        ws_k.cell(row=ri, column=7, value=round(r.qty1, 6))
        ws_k.cell(row=ri, column=8, value=round(r.qty2, 6))
        ws_k.cell(row=ri, column=9, value=round(r.total, 6))
        if r.total < 0:
            for c in range(6, 10):
                ws_k.cell(row=ri, column=c).font = _MINHA_FONT
        ri += 1
    ws_k.cell(row=ri, column=8, value="TOPLAM").font = _HEADER_FONT
    ws_k.cell(row=ri, column=9, value=round(report.formwork_total_m2, 6)).font = _TOTAL_FONT
    ws_k.cell(row=ri, column=9).fill = _TOTAL_FILL

    # --- A  BETON
    ws_b = wb.create_sheet("A  BETON")
    ws_b.cell(row=1, column=7, value="ÇEVRE/ALAN")
    ws_b.cell(row=1, column=8, value="YÜKSEKLİK")
    ws_b.cell(row=1, column=9, value="TOPLAM")
    for c in (7, 8, 9):
        ws_b.cell(row=1, column=c).font = _HEADER_FONT
    ri = 2
    for r in report.concrete_rows:
        # Alt blok satirlari (GRO/KORUMA/CATI): GT'de H etiket, I/J/K miktarlar
        if r.category in ("GRO", "KORUMA", "ÇATI", "CATI"):
            ws_b.cell(row=ri, column=8, value=r.label)
            ws_b.cell(row=ri, column=9, value=round(r.qty1, 6))
            ws_b.cell(row=ri, column=10, value=round(r.qty2, 6))
            ws_b.cell(row=ri, column=11, value=round(r.total, 6))
        else:
            ws_b.cell(row=ri, column=6, value=r.label)
            ws_b.cell(row=ri, column=7, value=round(r.qty1, 6))
            ws_b.cell(row=ri, column=8, value=round(r.qty2, 6))
            ws_b.cell(row=ri, column=9, value=round(r.total, 6))
        ri += 1

    wb.save(output_path)
    logger.info("Kumluca-layout Excel yazildi: %s", output_path)
    return output_path
