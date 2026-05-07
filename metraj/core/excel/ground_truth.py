"""Reference / ground-truth Excel reader.

Used both by Faz 0 PoC validation and by the revision comparator to
extract the per-room area / perimeter / minha values from the firma's
existing manually maintained workbook.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class GroundTruthRoom:
    code: str
    name: str
    floor: str
    floor_tip: str
    wall_tip: str
    ceiling_tip: str
    skirting_poz: str
    area: float
    perimeter: float
    height: float
    floor_minha: float = 0.0
    ceiling_minha: float = 0.0
    skirting_minha: float = 0.0
    wall_minha: float = 0.0


@dataclass
class GroundTruthBook:
    rooms: List[GroundTruthRoom] = field(default_factory=list)
    totals_by_kategori: Dict[str, float] = field(default_factory=dict)

    def by_code(self) -> Dict[str, GroundTruthRoom]:
        return {r.code: r for r in self.rooms if r.code}


class GroundTruthReader:
    """Parse the reference ``YM-K-Amfi Mahal Metraj_R22.xlsx`` style workbook.

    The reader is tolerant: missing columns degrade to ``0.0`` and missing
    rooms are simply skipped, never aborting the load.
    """

    def __init__(
        self,
        mahal_sheet_name: str = "Kutuphane MAHAL",
        icmal_sheet_name: str = "Icmal",
    ) -> None:
        self.mahal_sheet_name = mahal_sheet_name
        self.icmal_sheet_name = icmal_sheet_name

    def read(self, path: str | Path) -> GroundTruthBook:
        wb = load_workbook(path, data_only=True, read_only=True)
        rooms = self._read_rooms(wb)
        kategori_totals = self._read_icmal(wb)
        return GroundTruthBook(rooms=rooms, totals_by_kategori=kategori_totals)

    def _read_rooms(self, wb) -> List[GroundTruthRoom]:
        # Try the configured name first, then fuzzy match.
        sheet_name = self._resolve_sheet(wb, [self.mahal_sheet_name, "Mahal", "MAHAL"])
        if not sheet_name:
            return []
        ws = wb[sheet_name]
        rooms: List[GroundTruthRoom] = []
        # Column layout follows the reference workbook: row 8 is the header.
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not row or len(row) < 16:
                continue
            floor = str(row[2] or "").strip()
            floor_tip = str(row[3] or "").strip()
            skirting = str(row[4] or "").strip()
            wall_tip = str(row[5] or "").strip()
            ceiling_tip = str(row[6] or "").strip()
            code = str(row[8] or "").strip()
            name = str(row[9] or "").strip()
            if not code:
                continue
            try:
                area = float(row[12] or 0.0)
                perimeter = float(row[13] or 0.0)
                height = float(row[14] or 0.0)
            except (TypeError, ValueError):
                continue
            try:
                floor_minha = float(row[16] or 0.0)
                ceiling_minha = float(row[17] or 0.0)
                skirting_minha = float(row[20] or 0.0)
                wall_minha = float(row[21] or 0.0)
            except (TypeError, ValueError):
                floor_minha = ceiling_minha = skirting_minha = wall_minha = 0.0
            rooms.append(
                GroundTruthRoom(
                    code=code,
                    name=name,
                    floor=floor,
                    floor_tip=floor_tip,
                    wall_tip=wall_tip,
                    ceiling_tip=ceiling_tip,
                    skirting_poz=skirting,
                    area=area,
                    perimeter=perimeter,
                    height=height,
                    floor_minha=floor_minha,
                    ceiling_minha=ceiling_minha,
                    skirting_minha=skirting_minha,
                    wall_minha=wall_minha,
                )
            )
        return rooms

    def _read_icmal(self, wb) -> Dict[str, float]:
        totals: Dict[str, float] = {}
        sheet_name = self._resolve_sheet(wb, ["Icmal", "İcmal"])
        if not sheet_name:
            return totals
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or len(row) < 13:
                continue
            level = row[0]
            kategori = str(row[1] or row[2] or "").strip()
            if level == 2 and kategori:
                try:
                    totals[kategori] = float(row[13] or 0.0)
                except (TypeError, ValueError):
                    continue
        return totals

    @staticmethod
    def _resolve_sheet(wb, candidates: Iterable[str]) -> Optional[str]:
        existing = {name.lower(): name for name in wb.sheetnames}
        for cand in candidates:
            if cand.lower() in existing:
                return existing[cand.lower()]
        # Fuzzy contains
        for sheet in wb.sheetnames:
            for cand in candidates:
                if cand.lower() in sheet.lower():
                    return sheet
        return None
