"""Excel report writer.

Produces a workbook with the same conceptual layout as the reference
``YM-K-Amfi Mahal Metraj_R22.xlsx``:

* ``Kutuphane MAHAL`` -- per-room metraj sheet (mahal listesi)
* ``minha`` -- door/window deductions per floor
* ``Icmal`` -- poz aggregation with unit price totals

The format is *compatible* (same column headings) but not pixel-identical to
the reference; downstream firma can copy-paste cells into the official
template if branding is required.  The R22 ``#REF!`` formula errors of the
reference are *not* reproduced -- our writer outputs concrete values.
"""
from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..openings import Opening
from ..pozlar import IcmalRow, PozTotals
from ..quantities import RoomQuantities

logger = logging.getLogger(__name__)


_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
_TOTAL_FILL = PatternFill("solid", fgColor="305496")
_CATEGORY_FILL = PatternFill("solid", fgColor="FFE699")


class ExcelReportWriter:
    """Compose and write the metraj workbook."""

    def write(
        self,
        path: str | Path,
        quantities: Sequence[RoomQuantities],
        openings: Sequence[Opening],
        icmal: PozTotals,
        project_name: str = "",
    ) -> Path:
        wb = Workbook()
        # default sheet -> Kutuphane MAHAL
        ws_mahal = wb.active
        ws_mahal.title = "Kutuphane MAHAL"
        self._write_mahal_sheet(ws_mahal, quantities, project_name)

        ws_minha = wb.create_sheet("minha")
        self._write_minha_sheet(ws_minha, openings)

        ws_icmal = wb.create_sheet("Icmal")
        self._write_icmal_sheet(ws_icmal, icmal, project_name)

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        logger.info("Excel report written to %s", out)
        return out

    def _write_mahal_sheet(
        self,
        ws,
        quantities: Sequence[RoomQuantities],
        project_name: str,
    ) -> None:
        if project_name:
            ws.cell(row=1, column=1, value=project_name).font = Font(bold=True, size=14)
        headers = [
            "TIP", "KAT", "DOSEME KODU", "SUPURGELIK KODU", "DUVAR KODU",
            "TAVAN KODU", "MAHAL KODU", "MAHAL ADI",
            "ADET", "ALAN (M2)", "CEVRE (M)", "YUKSEKLIK (M)",
            "DOSEME MINHA (M2)", "TAVAN MINHA (M2)",
            "SUPURGELIK MINHA (M)", "DUVAR MINHA EKSTRA (M2)",
            "KAPI ADET", "PENCERE ADET",
            "TOPLAM KAPI ALANI (M2)", "TOPLAM PENCERE ALANI (M2)",
            "NET DOSEME (M2)", "NET SUPURGELIK (M)",
            "NET IC DUVAR (M2)", "NET TAVAN (M2)",
        ]
        header_row = 3
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=header)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_idx = header_row + 1
        for q in quantities:
            r = q.room
            values = [
                "B" if r.floor.upper() == "B" else r.floor,
                r.floor,
                r.floor_tip or "",
                r.skirting_tip or "",
                r.wall_tip or "",
                r.ceiling_tip or "",
                r.code,
                r.name,
                1,
                round(r.area, 3),
                round(r.perimeter, 3),
                round(r.height, 3),
                q.floor_minha_m2,
                q.ceiling_minha_m2,
                q.skirting_minha_m,
                q.wall_minha_extra_m2,
                q.door_count,
                q.window_count,
                q.door_minha_m2,
                q.window_minha_m2,
                q.net_floor_m2,
                q.net_skirting_m,
                q.net_wall_m2,
                q.net_ceiling_m2,
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            row_idx += 1

        # Total row
        total_row = row_idx + 1
        ws.cell(row=total_row, column=8, value="GENEL TOPLAM").font = _TOTAL_FONT
        for col in (10, 11, 21, 22, 23, 24):
            letter = get_column_letter(col)
            ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({letter}{header_row + 1}:{letter}{row_idx - 1})",
            ).font = _TOTAL_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=col).fill = _TOTAL_FILL

        # Auto column widths
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(28, len(header) + 2))
        ws.freeze_panes = "I4"

    def _write_minha_sheet(self, ws, openings: Sequence[Opening]) -> None:
        ws.cell(row=1, column=1, value="MINHA - Kapi/Pencere Dusumleri").font = Font(bold=True, size=12)
        headers = ["KAT", "TUR", "OLCU (cm/cm)", "ADET", "EN (m)", "YUKSEKLIK (m)", "TOPLAM (m2)"]
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col_idx, value=header)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL

        groups: Dict[Tuple[str, str, str], List[Opening]] = defaultdict(list)
        for opening in openings:
            if not opening.floor:
                continue
            groups[(opening.floor, opening.kind, opening.dim_label)].append(opening)
        row_idx = 4
        for (floor, kind, dim), group in sorted(groups.items()):
            count = len(group)
            sample = group[0]
            total = sum(o.area for o in group)
            ws.cell(row=row_idx, column=1, value=floor)
            ws.cell(row=row_idx, column=2, value=kind)
            ws.cell(row=row_idx, column=3, value=dim)
            ws.cell(row=row_idx, column=4, value=count)
            ws.cell(row=row_idx, column=5, value=round(sample.width_m, 3))
            ws.cell(row=row_idx, column=6, value=round(sample.height_m, 3))
            ws.cell(row=row_idx, column=7, value=round(total, 3))
            row_idx += 1

        total_row = row_idx + 1
        ws.cell(row=total_row, column=1, value="TOPLAM").font = _TOTAL_FONT
        ws.cell(
            row=total_row,
            column=7,
            value=f"=SUM(G4:G{row_idx - 1})",
        ).font = _TOTAL_FONT

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

    def _write_icmal_sheet(self, ws, totals: PozTotals, project_name: str) -> None:
        if project_name:
            ws.cell(row=1, column=1, value=project_name).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value="ICMAL - Insaat Isleri Kesfi").font = Font(bold=True, size=12)
        headers = ["KATEGORI", "POZ NO", "TANIM", "BIRIM", "MIKTAR",
                   "BIRIM FIYAT (TL)", "TUTAR (TL)"]
        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=header)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL

        row_idx = header_row + 1
        current_kategori = None
        kategori_subtotal_rows: List[Tuple[str, int]] = []
        kategori_start_row = row_idx
        for entry in totals.rows:
            if current_kategori != entry.kategori:
                if current_kategori is not None:
                    self._write_subtotal(ws, current_kategori, kategori_start_row, row_idx - 1)
                    row_idx += 1
                current_kategori = entry.kategori
                ws.cell(row=row_idx, column=1, value=entry.kategori).font = _HEADER_FONT
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = _CATEGORY_FILL
                row_idx += 1
                kategori_start_row = row_idx
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value=entry.poz_no)
            ws.cell(row=row_idx, column=3, value=entry.tanim)
            ws.cell(row=row_idx, column=4, value=entry.birim)
            ws.cell(row=row_idx, column=5, value=entry.miktar)
            ws.cell(row=row_idx, column=6, value=entry.birim_fiyat)
            ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*F{row_idx}")
            row_idx += 1
        if current_kategori is not None:
            self._write_subtotal(ws, current_kategori, kategori_start_row, row_idx - 1)
            row_idx += 1

        # Grand total
        ws.cell(row=row_idx + 1, column=3, value="GENEL TOPLAM (TL)").font = _TOTAL_FONT
        ws.cell(
            row=row_idx + 1,
            column=7,
            value=f"=SUM(G{header_row + 1}:G{row_idx - 1})",
        ).font = _TOTAL_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx + 1, column=col).fill = _TOTAL_FILL

        widths = [16, 18, 60, 8, 14, 18, 18]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = f"A{header_row + 1}"

    @staticmethod
    def _write_subtotal(ws, kategori: str, start_row: int, end_row: int) -> None:
        if end_row < start_row:
            return
        sub_row = end_row + 1
        ws.cell(row=sub_row, column=3, value=f"{kategori} ARA TOPLAM").font = Font(bold=True)
        ws.cell(row=sub_row, column=7,
                value=f"=SUM(G{start_row}:G{end_row})").font = Font(bold=True)
